import pandas as pd
import openpyxl
from openpyxl import load_workbook
import csv
import json
import xml.etree.ElementTree as ET
from pathlib import Path
import logging
import time
import gc
import re
import tracemalloc
from typing import Dict, List, Optional, Union, Iterator, Tuple
from io import StringIO, BytesIO
import gzip
import zipfile
import bz2

from .schemas import (
    ExcelParsingStats, 
    ExcelSecurityCheck, 
    OutputFormat, 
    CompressionType,
    ExcelWorkerConfig
)

logger = logging.getLogger(__name__)


class ExcelParserConfig:
    """Configuração do parser Excel"""
    
    def __init__(
        self,
        chunk_size: int = 50000,
        max_memory_mb: int = 2048,
        enable_streaming: bool = True,
        normalize_columns: bool = True,
        remove_empty_rows: bool = True,
        date_format: Optional[str] = None,
        decimal_separator: str = "."
    ):
        self.chunk_size = chunk_size
        self.max_memory_mb = max_memory_mb
        self.enable_streaming = enable_streaming
        self.normalize_columns = normalize_columns
        self.remove_empty_rows = remove_empty_rows
        self.date_format = date_format
        self.decimal_separator = decimal_separator


class ExcelSecurityValidator:
    """Validador de segurança para arquivos Excel"""
    
    MACRO_EXTENSIONS = {'.xlsm', '.xltm', '.xlam'}
    SAFE_EXTENSIONS = {'.xlsx', '.xls', '.csv', '.tsv'}
    
    @classmethod
    def check_file_security(cls, file_path: Path) -> ExcelSecurityCheck:
        """Verifica segurança do arquivo Excel"""
        
        try:
            file_extension = file_path.suffix.lower()
            filename = file_path.name
            
            # Verificar extensão
            is_macro_enabled = file_extension in cls.MACRO_EXTENSIONS
            has_vba_code = False
            has_external_references = False
            
            # Para arquivos Excel, fazer verificação mais profunda
            if file_extension in {'.xlsx', '.xlsm', '.xls'}:
                try:
                    workbook = load_workbook(file_path, data_only=True, read_only=True)
                    
                    # Verificar referências externas
                    if hasattr(workbook, 'defined_names'):
                        for name in workbook.defined_names:
                            if name.value and ('http://' in name.value or 'https://' in name.value):
                                has_external_references = True
                                break
                    
                    # Verificar VBA (aproximado)
                    if hasattr(workbook, 'vba_archive') and workbook.vba_archive:
                        has_vba_code = True
                        
                    workbook.close()
                    
                except Exception as e:
                    logger.warning(f"Erro ao verificar arquivo {filename}: {e}")
            
            # Determinar nível de risco
            if is_macro_enabled or has_vba_code:
                risk_level = "high"
                blocked_reason = "Arquivo contém macros ou código VBA"
                allowed = False
            elif has_external_references:
                risk_level = "medium" 
                blocked_reason = None
                allowed = True
            else:
                risk_level = "low"
                blocked_reason = None
                allowed = True
            
            return ExcelSecurityCheck(
                filename=filename,
                file_extension=file_extension,
                is_macro_enabled=is_macro_enabled,
                has_external_references=has_external_references,
                has_vba_code=has_vba_code,
                security_risk_level=risk_level,
                blocked_reason=blocked_reason,
                allowed_to_process=allowed
            )
            
        except Exception as e:
            logger.error(f"Erro na verificação de segurança: {e}")
            return ExcelSecurityCheck(
                filename=filename,
                file_extension=file_path.suffix,
                is_macro_enabled=True,  # Assumir o pior em caso de erro
                has_external_references=True,
                has_vba_code=True,
                security_risk_level="high",
                blocked_reason=f"Erro na verificação: {str(e)}",
                allowed_to_process=False
            )


class ExcelStreamProcessor:
    """Processador de Excel com streaming para arquivos grandes"""
    
    def __init__(self, config: ExcelParserConfig):
        self.config = config
        self.stats = ExcelParsingStats(
            total_sheets=0,
            sheets_processed=0,
            total_rows_read=0,
            total_rows_written=0,
            empty_rows_skipped=0,
            empty_columns_removed=0,
            processing_time_seconds=0.0,
            memory_peak_mb=0.0,
            chunk_count=0
        )
    
    def normalize_column_name(self, name: str) -> str:
        """Normaliza nome de coluna"""
        if not self.config.normalize_columns:
            return name
            
        # Remover caracteres especiais e espaços
        normalized = re.sub(r'[^\w\s]', '', str(name))
        normalized = re.sub(r'\s+', '_', normalized.strip())
        normalized = normalized.lower()
        
        # Garantir que não seja vazio
        if not normalized:
            return 'col_unnamed'
            
        return normalized
    
    def read_excel_streaming(
        self, 
        file_path: Path, 
        sheet_name: Optional[str] = None
    ) -> Iterator[pd.DataFrame]:
        """Lê Excel em chunks para economizar memória"""
        
        try:
            # Para CSV, usar pandas diretamente
            if file_path.suffix.lower() == '.csv':
                yield from pd.read_csv(
                    file_path,
                    chunksize=self.config.chunk_size,
                    encoding='utf-8',
                    on_bad_lines='skip'
                )
                return
            
            # Para Excel, tentar streaming com openpyxl
            workbook = load_workbook(file_path, data_only=True, read_only=True)
            
            sheet_names = [sheet_name] if sheet_name else workbook.sheetnames
            
            for sheet_name in sheet_names:
                worksheet = workbook[sheet_name]
                
                # Ler dados em chunks
                rows_data = []
                header_row = None
                row_count = 0
                
                for row in worksheet.iter_rows(values_only=True):
                    if header_row is None:
                        # Primeira linha como header
                        header_row = [self.normalize_column_name(cell) for cell in row if cell is not None]
                        continue
                    
                    # Filtrar linhas vazias se configurado
                    if self.config.remove_empty_rows and all(cell is None or cell == '' for cell in row):
                        self.stats.empty_rows_skipped += 1
                        continue
                    
                    rows_data.append(row)
                    row_count += 1
                    
                    # Yield chunk quando atingir tamanho
                    if len(rows_data) >= self.config.chunk_size:
                        df = pd.DataFrame(rows_data, columns=header_row)
                        yield df
                        
                        rows_data = []
                        self.stats.chunk_count += 1
                        self.stats.total_rows_read += len(df)
                        
                        # Limpar memória
                        del df
                        gc.collect()
                
                # Yield último chunk se houver dados restantes
                if rows_data:
                    df = pd.DataFrame(rows_data, columns=header_row)
                    yield df
                    self.stats.chunk_count += 1
                    self.stats.total_rows_read += len(df)
            
            workbook.close()
            
        except Exception as e:
            logger.error(f"Erro ao ler Excel em streaming: {e}")
            raise
    
    def process_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Processa DataFrame aplicando configurações"""
        
        # Remover colunas vazias se configurado
        if self.config.remove_empty_rows:
            before_cols = len(df.columns)
            df = df.dropna(axis=1, how='all')
            self.stats.empty_columns_removed += before_cols - len(df.columns)
        
        # Converter tipos de data se especificado
        if self.config.date_format:
            for col in df.columns:
                if df[col].dtype == 'object':
                    try:
                        df[col] = pd.to_datetime(df[col], format=self.config.date_format, errors='ignore')
                    except:
                        pass
        
        # Ajustar separador decimal
        if self.config.decimal_separator == ',':
            numeric_cols = df.select_dtypes(include=[float, int]).columns
            for col in numeric_cols:
                df[col] = df[col].astype(str).str.replace('.', ',')
        
        return df


class ExcelFormatWriter:
    """Escritor para diferentes formatos de saída"""
    
    def __init__(self, output_format: OutputFormat, compression: CompressionType):
        self.output_format = output_format
        self.compression = compression
    
    def write_dataframe(
        self, 
        df: pd.DataFrame, 
        output_path: Path,
        sheet_name: str = "Sheet1"
    ):
        """Escreve DataFrame no formato especificado"""
        
        try:
            if self.output_format == OutputFormat.CSV:
                self._write_csv(df, output_path)
            elif self.output_format == OutputFormat.JSON:
                self._write_json(df, output_path)
            elif self.output_format == OutputFormat.XML:
                self._write_xml(df, output_path, sheet_name)
            elif self.output_format == OutputFormat.TSV:
                self._write_tsv(df, output_path)
            elif self.output_format == OutputFormat.PARQUET:
                self._write_parquet(df, output_path)
            else:
                raise ValueError(f"Formato não suportado: {self.output_format}")
                
        except Exception as e:
            logger.error(f"Erro ao escrever formato {self.output_format}: {e}")
            raise
    
    def _write_csv(self, df: pd.DataFrame, output_path: Path):
        """Escreve CSV"""
        df.to_csv(output_path, index=False, encoding='utf-8')
        self._apply_compression(output_path)
    
    def _write_json(self, df: pd.DataFrame, output_path: Path):
        """Escreve JSON"""
        df.to_json(output_path, orient='records', indent=2, force_ascii=False)
        self._apply_compression(output_path)
    
    def _write_xml(self, df: pd.DataFrame, output_path: Path, sheet_name: str):
        """Escreve XML"""
        root = ET.Element("workbook")
        sheet_elem = ET.SubElement(root, "sheet", name=sheet_name)
        
        for _, row in df.iterrows():
            row_elem = ET.SubElement(sheet_elem, "row")
            for col, value in row.items():
                col_elem = ET.SubElement(row_elem, "cell", column=str(col))
                col_elem.text = str(value) if pd.notna(value) else ""
        
        tree = ET.ElementTree(root)
        tree.write(output_path, encoding='utf-8', xml_declaration=True)
        self._apply_compression(output_path)
    
    def _write_tsv(self, df: pd.DataFrame, output_path: Path):
        """Escreve TSV"""
        df.to_csv(output_path, index=False, sep='\t', encoding='utf-8')
        self._apply_compression(output_path)
    
    def _write_parquet(self, df: pd.DataFrame, output_path: Path):
        """Escreve Parquet"""
        df.to_parquet(output_path, index=False)
        self._apply_compression(output_path)
    
    def _apply_compression(self, file_path: Path):
        """Aplica compressão se especificada"""
        if self.compression == CompressionType.NONE:
            return
        
        compressed_path = file_path.with_suffix(f"{file_path.suffix}.{self.compression.value}")
        
        with open(file_path, 'rb') as f_in:
            if self.compression == CompressionType.GZIP:
                with gzip.open(compressed_path, 'wb') as f_out:
                    f_out.write(f_in.read())
            elif self.compression == CompressionType.ZIP:
                with zipfile.ZipFile(compressed_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                    zf.write(file_path, file_path.name)
            elif self.compression == CompressionType.BZIP2:
                with bz2.open(compressed_path, 'wb') as f_out:
                    f_out.write(f_in.read())
        
        # Remover arquivo original
        file_path.unlink()
        
        return compressed_path


def parse_excel_to_format(
    input_path: Path,
    output_path: Path,
    output_format: OutputFormat,
    config: ExcelParserConfig,
    sheets: Optional[List[str]] = None,
    compression: CompressionType = CompressionType.NONE
) -> ExcelParsingStats:
    """
    Função principal para converter Excel para outros formatos
    
    Args:
        input_path: Caminho do arquivo Excel de entrada
        output_path: Caminho do arquivo de saída
        output_format: Formato de saída desejado
        config: Configurações do parser
        sheets: Lista de planilhas específicas (se None, processa todas)
        compression: Tipo de compressão
    
    Returns:
        Estatísticas do processamento
    """
    
    start_time = time.time()
    tracemalloc.start()
    
    try:
        # Verificar segurança do arquivo
        security_check = ExcelSecurityValidator.check_file_security(input_path)
        if not security_check.allowed_to_process:
            raise ValueError(f"Arquivo bloqueado: {security_check.blocked_reason}")
        
        # Inicializar processador e escritor
        processor = ExcelStreamProcessor(config)
        writer = ExcelFormatWriter(output_format, compression)
        
        # Determinar planilhas a processar
        if input_path.suffix.lower() == '.csv':
            sheet_names = ['Sheet1']  # CSV tem apenas uma "planilha"
        else:
            workbook = load_workbook(input_path, data_only=True, read_only=True)
            all_sheets = workbook.sheetnames
            sheet_names = sheets if sheets else all_sheets
            workbook.close()
        
        processor.stats.total_sheets = len(sheet_names)
        
        # Processar cada planilha
        for sheet_name in sheet_names:
            logger.info(f"Processando planilha: {sheet_name}")
            
            # Determinar caminho de saída para esta planilha
            if len(sheet_names) > 1:
                sheet_output_path = output_path.parent / f"{output_path.stem}_{sheet_name}{output_path.suffix}"
            else:
                sheet_output_path = output_path
            
            # Processar em chunks
            all_chunks = []
            
            for chunk_df in processor.read_excel_streaming(input_path, sheet_name):
                processed_chunk = processor.process_dataframe(chunk_df)
                all_chunks.append(processed_chunk)
                
                processor.stats.total_rows_written += len(processed_chunk)
                
                # Monitor de memória
                current, peak = tracemalloc.get_traced_memory()
                processor.stats.memory_peak_mb = max(
                    processor.stats.memory_peak_mb, 
                    peak / 1024 / 1024
                )
            
            # Combinar chunks e escrever
            if all_chunks:
                final_df = pd.concat(all_chunks, ignore_index=True)
                writer.write_dataframe(final_df, sheet_output_path, sheet_name)
                
                # Limpar memória
                del final_df, all_chunks
                gc.collect()
            
            processor.stats.sheets_processed += 1
        
        # Finalizar estatísticas
        processor.stats.processing_time_seconds = time.time() - start_time
        
        logger.info(f"Conversão concluída: {processor.stats.sheets_processed} planilhas, "
                   f"{processor.stats.total_rows_written} linhas, "
                   f"{processor.stats.processing_time_seconds:.2f}s")
        
        return processor.stats
        
    except Exception as e:
        logger.error(f"Erro na conversão: {e}")
        raise
    finally:
        tracemalloc.stop()


def get_excel_info(file_path: Path) -> Dict:
    """Extrai informações básicas do arquivo Excel"""
    
    try:
        if file_path.suffix.lower() == '.csv':
            # Para CSV, contar linhas rapidamente
            with open(file_path, 'r', encoding='utf-8') as f:
                row_count = sum(1 for line in f) - 1  # -1 para header
            
            return {
                "sheets_count": 1,
                "sheets_names": ["Sheet1"],
                "total_rows": row_count,
                "total_columns": 0,  # Seria necessário ler primeira linha
                "has_macros": False
            }
        
        # Para Excel
        workbook = load_workbook(file_path, data_only=True, read_only=True)
        
        sheets_names = workbook.sheetnames
        total_rows = 0
        total_columns = 0
        
        for sheet_name in sheets_names:
            worksheet = workbook[sheet_name]
            sheet_rows = worksheet.max_row
            sheet_cols = worksheet.max_column
            
            total_rows += sheet_rows
            total_columns = max(total_columns, sheet_cols)
        
        workbook.close()
        
        return {
            "sheets_count": len(sheets_names),
            "sheets_names": sheets_names,
            "total_rows": total_rows,
            "total_columns": total_columns,
            "has_macros": file_path.suffix.lower() in {'.xlsm', '.xltm'}
        }
        
    except Exception as e:
        logger.error(f"Erro ao obter info do Excel: {e}")
        return {
            "sheets_count": 0,
            "sheets_names": [],
            "total_rows": 0,
            "total_columns": 0,
            "has_macros": True  # Assumir o pior em caso de erro
        }